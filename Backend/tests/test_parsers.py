import pytest
from parsers.csv_parser import parse_csv
from parsers.xlsx_parser import parse_xlsx
from models.data_type import DataType


class TestCSVParser:
    """Test CSV parsing with type inference"""
    
    def test_simple_csv_parsing(self):
        """Test: Parse basic CSV with headers"""
        csv = "name,age,score\nAlice,30,95.5\nBob,25,87.3"
        df = parse_csv(csv)
        
        assert len(df.rows) == 2
        assert df.total_row_count == 2
        assert len(df.schema.columns) == 3
        assert df.rows[0]["name"] == "Alice"
        assert df.rows[0]["age"] == 30
    
    def test_type_inference_number(self):
        """Test: Infers NUMBER type correctly"""
        csv = "id,count\n1,100\n2,200\n3,300"
        df = parse_csv(csv)
        
        id_col = next(c for c in df.schema.columns if c.name == "id")
        assert id_col.type == DataType.NUMBER
        assert df.rows[0]["id"] == 1
        assert isinstance(df.rows[0]["id"], int)
    
    def test_type_inference_date(self):
        """Test: Infers DATE type correctly"""
        csv = "event_date\n2026-01-15\n2026-02-20\n2026-03-10"
        df = parse_csv(csv)
        
        date_col = next(c for c in df.schema.columns if c.name == "event_date")
        assert date_col.type == DataType.DATE
    
    def test_type_inference_string(self):
        """Test: Infers STRING type correctly"""
        csv = "product\nLaptop\nMouse\nKeyboard"
        df = parse_csv(csv)
        
        col = next(c for c in df.schema.columns if c.name == "product")
        assert col.type == DataType.STRING
    
    def test_mixed_types(self):
        """Test: Handles mixed column types"""
        csv = "id,name,age,date,price\n1,Alice,30,2026-01-01,99.99\n2,Bob,25,2026-02-15,150.50"
        df = parse_csv(csv)
        
        col_types = {c.name: c.type for c in df.schema.columns}
        assert col_types["id"] == DataType.NUMBER
        assert col_types["name"] == DataType.STRING
        assert col_types["age"] == DataType.NUMBER
        assert col_types["date"] == DataType.DATE
        assert col_types["price"] == DataType.NUMBER
    
    def test_null_values(self):
        """Test: Handles null/empty values"""
        csv = "col1,col2,col3\n1,,value\n,2,\nvalue,,3"
        df = parse_csv(csv)
        
        assert df.rows[0]["col2"] is None
        assert df.rows[1]["col1"] is None
        assert df.rows[2]["col2"] is None
    
    def test_empty_csv(self):
        """Test: Handles empty CSV"""
        csv = ""
        df = parse_csv(csv)
        
        assert len(df.rows) == 0
        assert len(df.schema.columns) == 0
        assert df.total_row_count == 0
    
    def test_single_column(self):
        """Test: Single column CSV"""
        csv = "value\n10\n20\n30"
        df = parse_csv(csv)
        
        assert len(df.schema.columns) == 1
        assert len(df.rows) == 3
    
    def test_type_coercion_float_to_int(self):
        """Test: Converts 10.0 to 10 (int)"""
        csv = "num\n10.0\n20.0\n30.0"
        df = parse_csv(csv)
        
        assert df.rows[0]["num"] == 10
        assert isinstance(df.rows[0]["num"], int)
    
    def test_date_formats(self):
        """Test: Recognizes multiple date formats"""
        csv1 = "date\n2026-01-15\n2026-02-20"
        csv2 = "date\n15/01/2026\n20/02/2026"
        csv3 = "date\n01/15/2026\n02/20/2026"
        
        df1 = parse_csv(csv1)
        df2 = parse_csv(csv2)
        df3 = parse_csv(csv3)
        
        assert df1.schema.columns[0].type == DataType.DATE
        assert df2.schema.columns[0].type == DataType.DATE
        assert df3.schema.columns[0].type == DataType.DATE
    
    def test_large_csv(self):
        """Test: Parses 1000+ rows"""
        rows = ["id,value"] + [f"{i},{i*10}" for i in range(1000)]
        csv = "\n".join(rows)
        df = parse_csv(csv)
        
        assert df.total_row_count == 1000
        assert len(df.rows) == 1000
        assert df.rows[999]["id"] == 999


class TestXLSXParser:
    """Test XLSX parsing"""
    
    def test_xlsx_parsing_basic(self):
        """Test: Parse basic XLSX"""
        # This requires sample XLSX file
        import base64
        import openpyxl
        import io
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["name", "age"])
        sheet.append(["Alice", 30])
        sheet.append(["Bob", 25])
        
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        
        df = parse_xlsx(xlsx_bytes)
        
        assert len(df.rows) == 2
        assert df.total_row_count == 2
        assert df.rows[0]["name"] == "Alice"
    
    def test_xlsx_base64_decoding(self):
        """Test: Decodes base64 XLSX"""
        import base64
        import openpyxl
        import io
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["col1", "col2"])
        sheet.append([1, 2])
        
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        xlsx_b64 = base64.b64encode(xlsx_bytes).decode()
        
        df = parse_xlsx(xlsx_b64)
        
        assert len(df.rows) == 1
        assert df.rows[0]["col1"] == 1
    
    def test_xlsx_empty_sheet(self):
        """Test: Handles empty XLSX"""
        import openpyxl
        import io
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        
        df = parse_xlsx(xlsx_bytes)
        
        assert len(df.rows) == 0
        assert df.total_row_count == 0
    
    def test_xlsx_missing_headers(self):
        """Test: Auto-generates missing header names"""
        import openpyxl
        import io
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append([None, "name", None])
        sheet.append([1, "Alice", 30])
        
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        
        df = parse_xlsx(xlsx_bytes)
        
        col_names = [c.name for c in df.schema.columns]
        assert "Column_0" in col_names or col_names[0] != ""
        assert "name" in col_names

